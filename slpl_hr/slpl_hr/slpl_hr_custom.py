import json
import frappe
import pprint
from frappe import _
from frappe.utils.response import build_response
import requests
import time as t
from datetime import *
import calendar
import re
import sys
import traceback
from datetime import datetime
from math import *
from frappe.auth import LoginManager
import frappe.utils.user
import openpyxl
from openpyxl import Workbook
# import mysql.connector as mariadb
# mariadb.connect(user='it@scantechlaser.com', password='Dinudins@123', database='1bd3e0294da19198')
# cursor = mariadb_connection.cursor()
# return cursor
# import MySQLdb
# conn=MySQLdb.connect(host='35.237.255.109',user='root',passwd='laser123')
# cursor = conn.cursor()

def validate_date(doc,method):
	getuser= frappe.db.sql("SELECT user_id FROM `tabEmployee` WHERE name='"+str(doc.employee)+"'",as_dict=True)
	if str(doc.leave_approver) == str(getuser[0]['user_id']):
		frappe.throw(_("Leave Approver and leave Applicant should not be same"))
	
	if str(doc.leave_approver) != str(frappe.session.user) and str(doc.status) != 'Open':
		frappe.throw(_("You Cannot Approve/Cancel this leave beacause you are not leave approver for this employee"))
#role = frappe.get_roles(frappe.session.user)

@frappe.whitelist(allow_guest=True)
def getItemDescriptionSales():
	mysql = "SELECT * FROM `tabSales Invoice Item`"
	getSales=frappe.db.sql(mysql, as_dict=True)
	if getSales:
		for i in getSales:
			frappe.msgprint(i.item_name)
			mysql="UPDATE `tabSales Invoice Item` SET description='"+str(i.item_name)+"' WHERE name='"+str(i.name)+"'"
			frappe.db.sql(mysql)


@frappe.whitelist(allow_guest=True)
def getStockEntryLedger(code=''):
	data = {}
	MYDATA = 0
	old_stock_value = 0.0
	old_stock_value_difference = 0.0
	new_stock_value = 0.0
	new_stock_value_difference = 0.0
	dataArray = []
	outArray = []
	valuaion_rate = 0
	balance_qty=0
	InQty = 0
	OutQty = 0
	first = False
	stock_value_difference = 0
	item_value = 0
	Rate=['Item']
	diagnose = []
	row_count = 0
	sheet_name = ''
	temp_rcount = 0
	cell_1 = ''
	count = 0
  	file_path = '/home/frappe/frappe-bench/sites/assets/itme_negative.xlsx'
  	f = []
  	if file_path:
		diagnose.append(file_path)
		wb=openpyxl.load_workbook(file_path)


		sheet_name=wb.get_sheet_names()
		# frappe.throw(_(sheet_name))
		sheet=wb.get_sheet_by_name("Sheet1")

		row_count=sheet.max_row
		column_count=sheet.max_column

		for i in range(1,row_count+1):
					# diagnose.append(i)
			cell_1=str((sheet.cell(row=i,column=1).value)).encode('utf-8').strip()
					
			if cell_1:
				temp_rcount=temp_rcount+1
				code = sheet.cell(row=i,column=1).value

				data = {}
				MYDATA = 0
				old_stock_value = 0.0
				old_stock_value_difference = 0.0
				new_stock_value = 0.0
				new_stock_value_difference = 0.0
				dataArray = []
				outArray = []
				valuaion_rate = 0
				balance_qty=0
				InQty = 0
				OutQty = 0
				first = False
				stock_value_difference = 0
				item_value = 0

				frappe.msgprint(code)

				mysql = "SELECT stock_value, stock_value_difference, actual_qty,incoming_rate, qty_after_transaction, valuation_rate, name, voucher_no FROM `tabStock Ledger Entry` WHERE item_code = '"+str(code)+"' and warehouse='Stores - SLPL' and posting_date BETWEEN CAST('2017-01-01' AS DATE) AND CAST('2019-04-30' AS DATE) ORDER BY posting_date, posting_time ASC"
				getEntry = frappe.db.sql(mysql, as_dict=True)
				if getEntry:

					for i in getEntry:

						MYDATA +=1
						if i.actual_qty>0:

							if MYDATA==1:
								item_value = i.incoming_rate * i.actual_qty
								valuaion_rate = item_value / i.qty_after_transaction

								balance_qty += i.actual_qty
								if "PREC-" in str(i.voucher_no):

									stock_value = valuaion_rate * i.qty_after_transaction
									stock_value_difference = i.incoming_rate

									mysql="UPDATE `tabStock Ledger Entry` SET valuation_rate='"+str(valuaion_rate)+"', stock_value='"+str(stock_value)+"',  stock_value_difference='"+str(stock_value_difference*i.actual_qty)+"' WHERE name='"+str(i.name)+"'"
									frappe.db.sql(mysql)
									dataArray.append(stock_value)

								first = True

							if MYDATA>1:

								if i.qty_after_transaction==0:

									valuaion_rate = 0

								else:
									if balance_qty ==0:
										valuaion_rate = i.incoming_rate
									else:
										valuaion_rate = (valuaion_rate+i.incoming_rate)/i.qty_after_transaction

									valuaion_rate = round(valuaion_rate,4)
									
									item_value = item_value + i.incoming_rate * i.actual_qty

									valuaion_rate = item_value / i.qty_after_transaction

									stock_value_difference = i.incoming_rate
									stock_value = round(valuaion_rate* i.qty_after_transaction, 4)

									# frappe.msgprint(stock_value)
									mysql="UPDATE `tabStock Ledger Entry` SET valuation_rate='"+str(valuaion_rate)+"', stock_value='"+str(stock_value)+"',  stock_value_difference='"+str(stock_value_difference*i.actual_qty)+"' WHERE name='"+str(i.name)+"'"
									frappe.db.sql(mysql)
									dataArray.append(stock_value)
								balance_qty += i.actual_qty
						else:

							# frappe.msgprint(i.voucher_no)
							item_value = item_value + (valuaion_rate) * i.actual_qty

							if i.qty_after_transaction==0:
								# frappe.msgprint(i.voucher_no)
								item_value = 0
								stock_value = valuaion_rate* i.qty_after_transaction
								valuaion_rate = round(valuaion_rate,4)
								outArray.append(stock_value)
								mysql="UPDATE `tabStock Ledger Entry` SET valuation_rate='"+str(valuaion_rate)+"', stock_value='"+str(stock_value)+"',  stock_value_difference='"+str(valuaion_rate*i.actual_qty)+"' WHERE name='"+str(i.name)+"'"
								frappe.db.sql(mysql)
								valuaion_rate=0
								balance_qty += i.actual_qty
							else:
								stock_value = valuaion_rate* i.qty_after_transaction
								valuaion_rate = round(valuaion_rate,4)
								# frappe.msgprint(stock_value)
								mysql="UPDATE `tabStock Ledger Entry` SET valuation_rate='"+str(valuaion_rate)+"', stock_value='"+str(item_value)+"',  stock_value_difference='"+str(valuaion_rate*i.actual_qty)+"' WHERE name='"+str(i.name)+"'"
								frappe.db.sql(mysql)
								# frappe.msgprint(stock_value)
								balance_qty += i.actual_qty
								outArray.append(stock_value)
				# f.append(i)
			else:
				frappe.local.response['data']=[]
				frappe.local.response['msg']="Please provide mandatory fields"
				frappe.local.response['status']=False
				return False

	######################################################################################

	# mysql = "SELECT stock_value, stock_value_difference, actual_qty,incoming_rate, qty_after_transaction, valuation_rate, name, voucher_no FROM `tabStock Ledger Entry` WHERE item_code = '"+str(code)+"' and warehouse='Stores - SLPL' and posting_date BETWEEN CAST('2017-01-01' AS DATE) AND CAST('2019-04-30' AS DATE) ORDER BY posting_date, posting_time ASC"
	# getEntry = frappe.db.sql(mysql, as_dict=True)
	# if getEntry:

	# 	for i in getEntry:

	# 		MYDATA +=1
	# 		if i.actual_qty>0:

	# 			if MYDATA==1:
	# 				item_value = i.incoming_rate * i.actual_qty
	# 				valuaion_rate = item_value / i.qty_after_transaction

	# 				balance_qty += i.actual_qty
	# 				if "PREC-" in str(i.voucher_no):

	# 					stock_value = valuaion_rate * i.qty_after_transaction
	# 					stock_value_difference = i.incoming_rate

	# 					mysql="UPDATE `tabStock Ledger Entry` SET valuation_rate='"+str(valuaion_rate)+"', stock_value='"+str(stock_value)+"',  stock_value_difference='"+str(stock_value_difference*i.actual_qty)+"' WHERE name='"+str(i.name)+"'"
	# 					frappe.db.sql(mysql)
	# 					dataArray.append(stock_value)

	# 				first = True

	# 			if MYDATA>1:

	# 				if i.qty_after_transaction==0:

	# 					valuaion_rate = 0

	# 				else:
	# 					if balance_qty ==0:
	# 						valuaion_rate = i.incoming_rate
	# 					else:
	# 						valuaion_rate = (valuaion_rate+i.incoming_rate)/i.qty_after_transaction

	# 					valuaion_rate = round(valuaion_rate,4)
						
	# 					item_value = item_value + i.incoming_rate * i.actual_qty

	# 					valuaion_rate = item_value / i.qty_after_transaction

	# 					stock_value_difference = i.incoming_rate
	# 					stock_value = round(valuaion_rate* i.qty_after_transaction, 4)

	# 					# frappe.msgprint(stock_value)
	# 					mysql="UPDATE `tabStock Ledger Entry` SET valuation_rate='"+str(valuaion_rate)+"', stock_value='"+str(stock_value)+"',  stock_value_difference='"+str(stock_value_difference*i.actual_qty)+"' WHERE name='"+str(i.name)+"'"
	# 					frappe.db.sql(mysql)
	# 					dataArray.append(stock_value)
	# 				balance_qty += i.actual_qty
	# 		else:

	# 			# frappe.msgprint(i.voucher_no)
	# 			item_value = item_value + (valuaion_rate) * i.actual_qty

	# 			if i.qty_after_transaction==0:
	# 				# frappe.msgprint(i.voucher_no)
	# 				item_value = 0
	# 				stock_value = valuaion_rate* i.qty_after_transaction
	# 				valuaion_rate = round(valuaion_rate,4)
	# 				outArray.append(stock_value)
	# 				mysql="UPDATE `tabStock Ledger Entry` SET valuation_rate='"+str(valuaion_rate)+"', stock_value='"+str(stock_value)+"',  stock_value_difference='"+str(valuaion_rate*i.actual_qty)+"' WHERE name='"+str(i.name)+"'"
	# 				frappe.db.sql(mysql)
	# 				valuaion_rate=0
	# 				balance_qty += i.actual_qty
	# 			else:
	# 				stock_value = valuaion_rate* i.qty_after_transaction
	# 				valuaion_rate = round(valuaion_rate,4)
	# 				# frappe.msgprint(stock_value)
	# 				mysql="UPDATE `tabStock Ledger Entry` SET valuation_rate='"+str(valuaion_rate)+"', stock_value='"+str(item_value)+"',  stock_value_difference='"+str(valuaion_rate*i.actual_qty)+"' WHERE name='"+str(i.name)+"'"
	# 				frappe.db.sql(mysql)
	# 				# frappe.msgprint(stock_value)
	# 				balance_qty += i.actual_qty
	# 				outArray.append(stock_value)

	# 		frappe.msgprint(str(i.voucher_no)+'---'+str(valuaion_rate)+'---'+str(i.incoming_rate)+'---'+str(i.qty_after_transaction)+'---'+str(item_value))

	# 	# frappe.msgprint("balance_qty"+str(balance_qty)+"   valuaion_rate"+str(valuaion_rate))
	# 	for i in dataArray:
	# 		InQty += i

	# 	for j in outArray:
	# 		OutQty += j

	# 	frappe.msgprint(dataArray)
	# 	frappe.msgprint(outArray)
	# 	frappe.msgprint(InQty)
	# 	frappe.msgprint(OutQty)

		####################################################################


#if 'HR Manager' not in role:

#from_date = str(doc.from_date).split("-")
#to_date = str(doc.from_date).split("-")
#posting_date = str(doc.posting_date).split("-")
#if posting_date[1] > to_date[1]:

#frappe.throw(_("You Cannot save Previous date Application, Please contact to HR Department"))
#else:

#frappe.msgprint("Continue")


@frappe.whitelist(allow_guest=True)
def Change_Negative(code=''):
	MYDATA = 0
	old_stock_value = 0.0
	old_stock_value_difference = 0.0
	new_stock_value = 0.0
	new_stock_value_difference = 0.0
	dataArray = []
	outArray = []
	valuaion_rate = 0
	balance_qty=0
	InQty = 0
	OutQty = 0
	first = False
	stock_value_difference = 0
	item_value = 0
	
	mysql = "SELECT stock_value, stock_value_difference, actual_qty,incoming_rate, qty_after_transaction, valuation_rate, name, voucher_no FROM `tabStock Ledger Entry` WHERE item_code = '"+str(code)+"' and warehouse='Stores - SLPL' and posting_date BETWEEN CAST('2017-01-01' AS DATE) AND CAST('2019-04-30' AS DATE) ORDER BY posting_date, posting_time ASC"
	getEntry = frappe.db.sql(mysql, as_dict=True)
	if getEntry:

		for i in getEntry:

			MYDATA +=1
			if i.actual_qty>0:

				if MYDATA==1:
					item_value = i.incoming_rate * i.actual_qty
					valuaion_rate = item_value / i.qty_after_transaction

					balance_qty += i.actual_qty
					if "PREC-" in str(i.voucher_no):

						stock_value = valuaion_rate * i.qty_after_transaction
						stock_value_difference = i.incoming_rate

						mysql="UPDATE `tabStock Ledger Entry` SET valuation_rate='"+str(valuaion_rate)+"', stock_value='"+str(stock_value)+"',  stock_value_difference='"+str(stock_value_difference*i.actual_qty)+"' WHERE name='"+str(i.name)+"'"
						frappe.db.sql(mysql)
						dataArray.append(stock_value)

					first = True

				if MYDATA>1:
					if i.qty_after_transaction==0:

						valuaion_rate = 0

					else:
						if balance_qty ==0:
							valuaion_rate = i.incoming_rate
						else:
							valuaion_rate = (valuaion_rate+i.incoming_rate)/i.qty_after_transaction

						valuaion_rate = round(valuaion_rate,4)

						frappe.msgprint(str(valuaion_rate)+'nikhil'+str(item_value))

						item_value = item_value + i.incoming_rate * i.actual_qty

						valuaion_rate = item_value / i.qty_after_transaction

						frappe.msgprint(str(valuaion_rate)+'sanket'+str(item_value))

						stock_value_difference = i.incoming_rate
						stock_value = round(valuaion_rate* i.qty_after_transaction, 4)

						# frappe.msgprint(stock_value)
						mysql="UPDATE `tabStock Ledger Entry` SET valuation_rate='"+str(valuaion_rate)+"', stock_value='"+str(stock_value)+"',  stock_value_difference='"+str(stock_value_difference*i.actual_qty)+"' WHERE name='"+str(i.name)+"'"
						frappe.db.sql(mysql)
						dataArray.append(stock_value)
					balance_qty += i.actual_qty
			else:

				# frappe.msgprint(i.voucher_no)
				item_value = item_value + (valuaion_rate) * i.actual_qty

				if i.qty_after_transaction==0:
					# frappe.msgprint(i.voucher_no)
					item_value = 0
					stock_value = valuaion_rate* i.qty_after_transaction
					valuaion_rate = round(valuaion_rate,4)
					outArray.append(stock_value)
					mysql="UPDATE `tabStock Ledger Entry` SET valuation_rate='"+str(valuaion_rate)+"', stock_value='"+str(stock_value)+"',  stock_value_difference='"+str(valuaion_rate*i.actual_qty)+"' WHERE name='"+str(i.name)+"'"
					frappe.db.sql(mysql)
					valuaion_rate=0
					balance_qty += i.actual_qty
				else:
					stock_value = valuaion_rate* i.qty_after_transaction
					valuaion_rate = round(valuaion_rate,4)
					# frappe.msgprint(stock_value)
					mysql="UPDATE `tabStock Ledger Entry` SET valuation_rate='"+str(valuaion_rate)+"', stock_value='"+str(item_value)+"',  stock_value_difference='"+str(valuaion_rate*i.actual_qty)+"' WHERE name='"+str(i.name)+"'"
					frappe.db.sql(mysql)
					# frappe.msgprint(stock_value)
					balance_qty += i.actual_qty
					outArray.append(stock_value)

			frappe.msgprint(str(i.voucher_no)+'---'+str(valuaion_rate)+'---'+str(i.incoming_rate)+'---'+str(i.qty_after_transaction)+'---'+str(item_value))
